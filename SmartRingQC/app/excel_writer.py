import openpyxl
from openpyxl import Workbook
import os

try:
    from app.paths import get_excel_path
except ImportError:
    from paths import get_excel_path

EXCEL_PATH = get_excel_path()
HEADERS = ['Captured on', 'Captured when', 'image_name', 'image_path', 'status', 'operator_id', 'batch_id', 'reject_reason', 'reviewed_at']

def init_excel():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspection Log"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)

def add_capture(timestamp, image_name, image_path):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    date_part = timestamp.split(' ')[0] if ' ' in timestamp else timestamp
    time_part = timestamp.split(' ')[1] if ' ' in timestamp else timestamp
    ws.append([date_part, time_part, image_name, image_path, 'WAITING', '', '', '', ''])
    wb.save(EXCEL_PATH)

def update_classification(image_name, status, operator_id, batch_id, reject_reason, reviewed_at):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    # Needs to match image_name which is in column 3 (C)
    # Search backwards since recent edits are likely at the bottom
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=3).value == image_name:
            ws.cell(row=row, column=5).value = status
            ws.cell(row=row, column=6).value = operator_id
            ws.cell(row=row, column=7).value = batch_id
            ws.cell(row=row, column=8).value = reject_reason or ""
            ws.cell(row=row, column=9).value = reviewed_at
            break
    wb.save(EXCEL_PATH)

def undo_classification(image_name):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=3).value == image_name:
            ws.cell(row=row, column=5).value = "WAITING"
            ws.cell(row=row, column=6).value = ""
            ws.cell(row=row, column=7).value = ""
            ws.cell(row=row, column=8).value = ""
            ws.cell(row=row, column=9).value = ""
            break
    wb.save(EXCEL_PATH)

def delete_classification(image_name):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=3).value == image_name:
            ws.delete_rows(row)
            break
    wb.save(EXCEL_PATH)

def generate_shift_report(stats, batch_data):
    from datetime import datetime
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    report_filename = f"Shift_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        from app.paths import get_reports_dir
    except ImportError:
        from paths import get_reports_dir
        
    path = os.path.join(get_reports_dir(), report_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Summary"
    
    ws.append(["Shift Report Summary", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws.append([])
    ws.append(["Total Captured", stats['total_captured']])
    ws.append(["Total Accepted", stats['accepted']])
    ws.append(["Total Rejected", stats['rejected']])
    ws.append(["Acceptance Rate %", f"{stats['rate']}%"])
    ws.append(["Most Common Reject Reason", stats['top_reject_reason']])
    ws.append([])
    
    ws.append(["Batch ID", "Accepted", "Rejected", "Total"])
    for b_id, b_acc, b_rej in batch_data:
        ws.append([b_id, b_acc, b_rej, (b_acc or 0) + (b_rej or 0)])
        
    # Auto-adjust columns width for better readability
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(path)
    return path
